#!/usr/bin/env python3
"""
Generate professional Excel reports for TRAORE GESTION PROJET
"""
import json
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_currency(amount):
    """Format number as currency"""
    return f"{amount:,.0f} CFA".replace(",", " ")

def create_excel_report(data, output_path):
    """Create Excel workbook with multiple sheets"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F4E79")
    subtitle_font = Font(bold=True, size=12, color="333333")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Colors for alternating rows
    row_even = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Parse data
    projects = data.get('projects', [])
    tasks = data.get('tasks', [])
    risks = data.get('risks', [])
    filters = data.get('filters', {})
    is_full = data.get('type') == 'full'
    
    include_tasks = is_full or filters.get('includeTasks', True)
    include_risks = is_full or filters.get('includeRisks', True)
    include_budget = is_full or filters.get('includeBudget', True)
    
    # ========== SHEET 1: RÉSUMÉ ==========
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    
    # Title
    ws_summary['B2'] = "RAPPORT DE GESTION DE PROJET"
    ws_summary['B2'].font = title_font
    ws_summary.merge_cells('B2:F2')
    
    ws_summary['B3'] = "TRAORE GESTION PROJET"
    ws_summary['B3'].font = subtitle_font
    
    ws_summary['B4'] = f"Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_summary['B5'] = f"Type: {'Rapport Complet' if is_full else 'Rapport Personnalisé'}"
    
    # Statistics
    stats_start = 7
    ws_summary[f'B{stats_start}'] = "STATISTIQUES GÉNÉRALES"
    ws_summary[f'B{stats_start}'].font = subtitle_font
    
    stats_data = [
        ["Métrique", "Valeur"],
        ["Nombre de projets", len(projects)],
        ["Projets actifs", len([p for p in projects if p.get('status') in ['Actif', 'En cours']])],
        ["Total tâches", len(tasks)],
        ["Tâches validées", len([t for t in tasks if t.get('status') == 'Validé'])],
        ["Tâches en cours", len([t for t in tasks if t.get('status') == 'En cours'])],
        ["Tâches en retard", len([t for t in tasks if t.get('status') == 'En retard'])],
        ["Total risques", len(risks)],
        ["Risques critiques", len([r for r in risks if r.get('severity') == 'Critique'])],
        ["Budget total (CFA)", sum([p.get('budgetPlanned', 0) for p in projects])],
        ["Total dépensé (CFA)", sum([p.get('budgetSpent', 0) for p in projects])],
    ]
    
    for i, row in enumerate(stats_data):
        for j, val in enumerate(row):
            cell = ws_summary.cell(row=stats_start + 1 + i, column=2 + j, value=val)
            cell.border = border
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            else:
                cell.alignment = left_align if j == 0 else center_align
                if i % 2 == 0:
                    cell.fill = row_even
    
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 20
    
    # ========== SHEET 2: PROJETS ==========
    ws_projects = wb.create_sheet("Projets")
    
    ws_projects['B2'] = "LISTE DES PROJETS"
    ws_projects['B2'].font = title_font
    
    project_headers = ["Nom", "Statut", "Responsable", "Budget prévu", "Dépenses", "Reste", "Tâches", "Date début", "Date fin"]
    
    for i, header in enumerate(project_headers):
        cell = ws_projects.cell(row=4, column=2 + i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    for row_idx, project in enumerate(projects):
        row_data = [
            project.get('name', ''),
            project.get('status', ''),
            project.get('responsibleName') or 'Non assigné',
            project.get('budgetPlanned', 0),
            project.get('budgetSpent', 0),
            project.get('budgetPlanned', 0) - project.get('budgetSpent', 0),
            len(project.get('tasks', [])),
            project.get('startDate', '')[:10] if project.get('startDate') else '-',
            project.get('endDate', '')[:10] if project.get('endDate') else '-',
        ]
        
        for col_idx, val in enumerate(row_data):
            cell = ws_projects.cell(row=5 + row_idx, column=2 + col_idx, value=val)
            cell.border = border
            cell.alignment = center_align if col_idx > 0 else left_align
            if row_idx % 2 == 1:
                cell.fill = row_even
    
    # Adjust column widths
    for i, width in enumerate([30, 12, 20, 15, 15, 15, 10, 12, 12]):
        ws_projects.column_dimensions[get_column_letter(2 + i)].width = width
    
    # ========== SHEET 3: TÂCHES ==========
    if include_tasks and tasks:
        ws_tasks = wb.create_sheet("Tâches")
        
        ws_tasks['B2'] = "LISTE DES TÂCHES"
        ws_tasks['B2'].font = title_font
        
        task_headers = ["Titre", "Statut", "Priorité", "Projet", "Assigné", "Date limite"]
        
        for i, header in enumerate(task_headers):
            cell = ws_tasks.cell(row=4, column=2 + i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        for row_idx, task in enumerate(tasks):
            project_name = ''
            if task.get('project'):
                project_name = task['project'].get('name', '')
            
            row_data = [
                task.get('title', ''),
                task.get('status', ''),
                task.get('priority', ''),
                project_name,
                task.get('assigneeName') or 'Non assigné',
                task.get('dueDate', '')[:10] if task.get('dueDate') else '-',
            ]
            
            for col_idx, val in enumerate(row_data):
                cell = ws_tasks.cell(row=5 + row_idx, column=2 + col_idx, value=val)
                cell.border = border
                cell.alignment = center_align if col_idx > 0 else left_align
                if row_idx % 2 == 1:
                    cell.fill = row_even
        
        for i, width in enumerate([40, 12, 12, 25, 20, 12]):
            ws_tasks.column_dimensions[get_column_letter(2 + i)].width = width
    
    # ========== SHEET 4: RISQUES ==========
    if include_risks and risks:
        ws_risks = wb.create_sheet("Risques")
        
        ws_risks['B2'] = "ANALYSE DES RISQUES"
        ws_risks['B2'].font = title_font
        
        risk_headers = ["Titre", "Sévérité", "Probabilité", "Statut", "Projet", "Mitigation"]
        
        for i, header in enumerate(risk_headers):
            cell = ws_risks.cell(row=4, column=2 + i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        for row_idx, risk in enumerate(risks):
            project_name = ''
            if risk.get('project'):
                project_name = risk['project'].get('name', '')
            
            row_data = [
                risk.get('title', ''),
                risk.get('severity', ''),
                risk.get('probability', ''),
                risk.get('status', ''),
                project_name,
                risk.get('mitigation') or '-',
            ]
            
            for col_idx, val in enumerate(row_data):
                cell = ws_risks.cell(row=5 + row_idx, column=2 + col_idx, value=val)
                cell.border = border
                cell.alignment = center_align if col_idx > 0 else left_align
                if row_idx % 2 == 1:
                    cell.fill = row_even
        
        for i, width in enumerate([30, 12, 12, 20, 25, 40]):
            ws_risks.column_dimensions[get_column_letter(2 + i)].width = width
    
    # ========== SHEET 5: BUDGET ==========
    if include_budget:
        ws_budget = wb.create_sheet("Budget")
        
        ws_budget['B2'] = "ANALYSE BUDGÉTAIRE"
        ws_budget['B2'].font = title_font
        
        budget_headers = ["Projet", "Budget prévu (CFA)", "Dépenses (CFA)", "Reste (CFA)", "% Utilisé", "Statut"]
        
        for i, header in enumerate(budget_headers):
            cell = ws_budget.cell(row=4, column=2 + i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        total_budget = 0
        total_spent = 0
        
        for row_idx, project in enumerate(projects):
            budget = project.get('budgetPlanned', 0)
            spent = project.get('budgetSpent', 0)
            percent = round((spent / budget) * 100) if budget > 0 else 0
            
            total_budget += budget
            total_spent += spent
            
            row_data = [
                project.get('name', ''),
                budget,
                spent,
                budget - spent,
                f"{percent}%",
                project.get('status', ''),
            ]
            
            for col_idx, val in enumerate(row_data):
                cell = ws_budget.cell(row=5 + row_idx, column=2 + col_idx, value=val)
                cell.border = border
                cell.alignment = center_align if col_idx > 0 else left_align
                if row_idx % 2 == 1:
                    cell.fill = row_even
        
        # Total row
        total_row = 5 + len(projects)
        total_percent = round((total_spent / total_budget) * 100) if total_budget > 0 else 0
        
        total_data = ["TOTAL", total_budget, total_spent, total_budget - total_spent, f"{total_percent}%", ""]
        for col_idx, val in enumerate(total_data):
            cell = ws_budget.cell(row=total_row, column=2 + col_idx, value=val)
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = center_align if col_idx > 0 else left_align
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        for i, width in enumerate([30, 18, 18, 18, 12, 12]):
            ws_budget.column_dimensions[get_column_letter(2 + i)].width = width
    
    # Save workbook
    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_excel.py <json_data_file> <output_path>")
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_path = sys.argv[2]
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    result = create_excel_report(data, output_path)
    print(f"Excel file created: {result}")
